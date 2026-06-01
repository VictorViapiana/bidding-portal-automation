import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

from openpyxl import load_workbook  
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

from src.config import STATUS_FILTROS, PALAVRAS_NEGATIVAS, PALAVRAS_POSITIVAS, obter_caminho_salvamento

def executar_raspagem():
    dados_recebendo_propostas = []
    dados_em_andamento = []
    
    opcoes = webdriver.ChromeOptions()
    
    # opcoes.add_argument("--headless=new") 
    
    opcoes.add_argument("--disable-gpu")
    opcoes.add_argument("--window-size=1920,1080")
    opcoes.add_argument("--start-maximized")
    
    navegador = webdriver.Chrome(options=opcoes)
    navegador.get("https://iam.secure.portaldecompraspublicas.com.br/realms/Portal/protocol/openid-connect/auth?client_id=aspclient&redirect_uri=https://operacao.portaldecompraspublicas.com.br/18/loginext/oAuth/&response_type=code&scope=openid")
    
    time.sleep(3)
    fechar_popup = navegador.find_element(By.ID, "fecharPopupLogin")
    fechar_popup.click()
    
    campo_login = navegador.find_element(By.ID, "username")
    campo_login.send_keys(os.getenv("PORTAL_USER"))
    
    campo_senha = navegador.find_element(By.ID, "password")
    campo_senha.send_keys(os.getenv("PORTAL_PASSWORD"))
    
    clicar_entrar = navegador.find_element(By.ID, "kc-login")
    clicar_entrar.click()
    
    time.sleep(5)
    
    try:
        aceitar_cookies = navegador.find_element(By.ID, "adopt-accept-all-button")
        navegador.execute_script("arguments[0].click();", aceitar_cookies)
    except Exception:
        pass
    
    clicar_processo = navegador.find_element(By.CLASS_NAME, "menuClose")
    clicar_processo.click()
    
    for status_atual in STATUS_FILTROS:
        print(f"\n================ INICIANDO FILTRO: {status_atual} ================\n", flush=True)
        
        clicar_pesquisa = navegador.find_element(By.XPATH, "//a[@title='Pesquisa de Processos']")
        clicar_pesquisa.click()
        time.sleep(3)
    
        objeto_1 = navegador.find_element(By.ID, "ttObjeto")
        objeto_1.send_keys("Software")
    
        campo_status = navegador.find_element(By.ID, "ttSTATUS")
        select_status = Select(campo_status)
        
        select_status.select_by_visible_text(status_atual)
    
        clicar_buscar = navegador.find_element(By.CLASS_NAME, "buttonDefault")
        clicar_buscar.click()
    
        print(
            f"{'='*35}\n"
            f"RELATÓRIO DE LICITAÇÕES\n\n"
            f"Status: {status_atual}\n"
            f"Objeto: Software\n"
            f"{'='*35}\n"
        )
    
        time.sleep(4)
        numero_pagina = 1
    
        while True:
            time.sleep(4)
            
            botoes_acao = navegador.find_elements(By.XPATH, "//table[@id='searchTableSorter']//a[@title='Visualizar Registro']")
            total_itens = len(botoes_acao)
            
            print(f"--- {status_atual} | Lendo a Página {numero_pagina} ({total_itens} itens encontrados) ---\n", flush=True)
            
            if total_itens == 0:
                print(f"Nenhum processo encontrado para o status {status_atual}.\n", flush=True)
                break
    
            for i in range(total_itens):
                print(f"Abrindo o processo {i + 1} de {total_itens} da Página {numero_pagina}...", flush=True)
                
                botoes_atualizados = navegador.find_elements(By.XPATH, "//table[@id='searchTableSorter']//a[@title='Visualizar Registro']")
                navegador.execute_script("arguments[0].click();", botoes_atualizados[i])
                
                time.sleep(3)
                
                try:
                    link_direto = navegador.current_url
                    
                    objeto = navegador.find_element(By.XPATH, "//p[b[contains(text(), 'Objeto:')]]")
                    descricao_detalhada = objeto.text.lower()
                    
                    texto_orgao = navegador.find_element(By.XPATH, "//p[b[contains(text(), 'Órgão:')]]").text
                    texto_unidade = navegador.find_element(By.XPATH, "//p[b[contains(text(), 'Modalidade')]]").text
    
                    orgao_puro = texto_orgao.replace("Órgão:", "").strip()
                    unidade_pura = texto_unidade.replace("Modalidade/Proc. Aux:", "").strip()
    
                    termos_positivos_encontrados = []
                    termos_negativos_encontrados = []
    
                    for termo in PALAVRAS_NEGATIVAS:
                        if termo in descricao_detalhada:
                            termos_negativos_encontrados.append(termo)
                            print(f"❌ Ignorado: Contém termo negativo desclassificatório: {termo}\n", flush=True)
                            break 
    
                    if not termos_negativos_encontrados:
                        for termo in PALAVRAS_POSITIVAS:
                            if termo in descricao_detalhada:
                                termos_positivos_encontrados.append(termo)
    
                        if termos_positivos_encontrados:
                            print("🎯 Aprovado!", flush=True)
                            print(f"Os termos {termos_positivos_encontrados} foram encontrados.\n", flush=True)
                            
                            info_licitacao = {
                                "Órgão": orgao_puro,
                                "Unidade de Compra": unidade_pura,
                                "Termos Encontrados": ", ".join(termos_positivos_encontrados),
                                "Link do Processo": link_direto
                            }
                            
                            if status_atual == "Recebendo Propostas":
                                dados_recebendo_propostas.append(info_licitacao)
                            else:
                                dados_em_andamento.append(info_licitacao)
                        else:
                            print("⚪ Ignorado: Objeto neutro (não possui termos relevantes para a Galactix).\n", flush=True)
                        
                except Exception as e:
                    print(f"⚠️ Não foi possível ler o campo Objeto. Erro: {e}", flush=True)
                
                navegador.back()
                time.sleep(3)
    
            numero_pagina += 1
            xpath_proxima_pagina = f"//a[@title='Página {numero_pagina}']"
            
            try:
                botao_proxima = navegador.find_element(By.XPATH, xpath_proxima_pagina)
                print(f"Avançando para a Página {numero_pagina}...", flush=True)
                navegador.execute_script("arguments[0].click();", botao_proxima)
            except:
                print(f"Chegamos ao fim das páginas de: {status_atual}.", flush=True)
                break
                
    time.sleep(2)
    navegador.quit()
    
    return dados_recebendo_propostas, dados_em_andamento

def salvar_e_formatar_planilha(dados_recebendo_propostas, dados_em_andamento):
    print("\n================ GERANDO PLANILHA EXCEL ================\n", flush=True)
    nome_planilha = obter_caminho_salvamento()
    
    with pd.ExcelWriter(nome_planilha, engine='openpyxl') as writer:
        df_recebendo = pd.DataFrame(dados_recebendo_propostas)
        df_andamento = pd.DataFrame(dados_em_andamento)
        
        if not df_recebendo.empty:
            df_recebendo.to_excel(writer, sheet_name="Recebendo Propostas", index=False)
        else:
            pd.DataFrame([{"Aviso": "Nenhuma licitação encontrada"}]).to_excel(writer, sheet_name="Recebendo Propostas", index=False)
            
        if not df_andamento.empty:
            df_andamento.to_excel(writer, sheet_name="Em Andamento", index=False)
        else:
            pd.DataFrame([{"Aviso": "Nenhuma licitação encontrada"}]).to_excel(writer, sheet_name="Em Andamento", index=False)
            
    workbook = load_workbook(nome_planilha)
    
    cor_cabecalho = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fonte_cabecalho = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fonte_dados = Font(name="Segoe UI", size=10, bold=False, color="000000")
    fonte_link = Font(name="Segoe UI", size=10, bold=False, color="0563C1", underline="single")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
    
    for nome_aba in workbook.sheetnames:
        aba = workbook[nome_aba]
        
        if aba["A1"].value == "Aviso":
            continue
            
        aba.views.sheetView[0].showGridLines = True
        
        aba.row_dimensions[1].height = 28
        for celula in aba[1]:
            celula.fill = cor_cabecalho
            celula.font = fonte_cabecalho
            celula.alignment = alinhamento_centro
            
        for linha in range(2, aba.max_row + 1):
            aba.row_dimensions[linha].height = 20
            
            for coluna in range(1, aba.max_column + 1):
                celula = aba.cell(row=linha, column=coluna)
                celula.font = fonte_dados
                celula.alignment = alinhamento_esquerda
                
                if aba.cell(row=1, column=coluna).value == "Link do Processo" and celula.value:
                    celula.hyperlink = celula.value
                    celula.font = fonte_link
    
        for col in aba.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            col_name = aba.cell(row=1, column=col[0].column).value
            
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value).strip()))
            
            if col_name == "Link do Processo":
                largura_calculada = min(max_len + 6, 120)
            else:
                largura_calculada = min(max_len + 4, 70)
                
            aba.column_dimensions[col_letter].width = largura_calculada if largura_calculada > 12 else 12
    
        ultima_coluna = get_column_letter(aba.max_column)
        aba.auto_filter.ref = f"A1:{ultima_coluna}{aba.max_row}"
    
    workbook.save(nome_planilha)
    print(f"📊 Sucesso!\nO arquivo '{nome_planilha}' foi gerado!", flush=True)